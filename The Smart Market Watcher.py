#The Smart Market Watcher
#Taif AlQarni & Bushra Alharthi
print("---Part |----")
import requests
from requests import HTTPError, Timeout, RequestException
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule

#The Decorator:
def execution_logger(func):
    def wrapper(self, symbol):
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] Fetching data for: {symbol}...")
        return func(self, symbol)
    return wrapper

#Input Handling:
with open('watchlist.txt', 'w', encoding='utf-8') as f:
    f.write('AAPL\n')
    f.write('TSLA\n')
    f.write('MSFT\n')
    f.write('NVDA\n')

def read_watchlist(filename='watchlist.txt'):
    with open(filename, 'r', encoding='utf-8') as f:
        symbols_list = [line.strip() for line in f if line.strip()]
    return symbols_list

#The API Client:
class StockClient:
    def __init__(self):
        self.headers = {"User-Agent": "Mozilla/5.0"}
        self.BASE_URL = "https://query1.finance.yahoo.com/v8/finance/chart/{SYMBOL}?interval=1d&range=2d"

    @execution_logger
    def fetchPrice(self, symbol):
        url = self.BASE_URL.format(SYMBOL=symbol)
        try:
            response = requests.get(url, headers=self.headers, timeout=10)
            response.raise_for_status()
            data = response.json()
            chart_result = data.get('chart', {}).get('result')

            meta = chart_result[0].get('meta', {})
            #Two keys to pull from json for calc
            current_price = meta.get('regularMarketPrice')
            previous_close = meta.get('chartPreviousClose')
            return current_price, previous_close

        except Timeout:
            print(f"Error in {symbol}: Request timed out.")
        except HTTPError as error:
            print(f"Error in {symbol}: HTTP error occurred: {error}")
        except RequestException as error:
            print(f"Error in {symbol}: Network error occurred: {error}")


symbolsList = read_watchlist()
print(f"Market Stocks to check: {symbolsList}\n")
client = StockClient()

#######################################
records = []
#Data Processing:
for symbol in symbolsList:
    current_price, previous_close = client.fetchPrice(symbol)

    if current_price is not None:
        if previous_close is None:
            previous_close = current_price

        change = round(current_price - previous_close, 2)
        if change > 0:
            status = "Positive"
        elif change < 0:
            status = "Negative"
        else:
            status = "No Change"

        print(f" Success: The current price for {symbol} is ${current_price:,.2f}\n")
    else:
        change = None
        status = "No Change"
        print(f" Failed to get price for {symbol}\n")

    records.append({
        'Symbol': symbol,
        'Current_Price$': current_price,
        'Previous_Close$': previous_close,
        'Status': status
    })

df = pd.DataFrame(records)
#calc the change% col
df['Change %'] = ((df['Current_Price$'] - df['Previous_Close$']) / df['Previous_Close$']) * 100
df['Change %'] = df['Change %'].round(2)

print("---Part ||----")
print(df)

#Excel Formatting:
today_date = datetime.now().strftime('%Y-%m-%d')
filename = f'Market_Report_{today_date}.xlsx'

df.to_excel(filename, index=False)

wb = load_workbook(filename)
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True)

#Adding the change col
change_col = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'Change %':
        change_col = col[0].column_letter
        break
#Visual Styling (Openpyxl):
if change_col:
    ws.conditional_formatting.add(
        f"{change_col}2:{change_col}{ws.max_row}",
        CellIsRule(operator='greaterThan', formula=['0'], font=Font(color="008000"))
    )
    ws.conditional_formatting.add(
        f"{change_col}2:{change_col}{ws.max_row}",
        CellIsRule(operator='lessThan', formula=['0'], font=Font(color="FF0000"))
    )

wb.save(filename)
print(f"\nExcel report created: {filename}")